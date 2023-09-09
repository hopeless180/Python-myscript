import openpyxl
from openpyxl import load_workbook

import openpyxl

# 打开Excel文件或新建一个工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 设置表头
worksheet.insert_rows(1)
worksheet.cell(row=1, column=1, value='作者名')
worksheet.cell(row=1, column=2, value='是否检查过')
worksheet.cell(row=1, column=3, value='备注')

# 打开txt文件，读取内容
with open(r'e:\作者.txt', 'r', encoding='utf-8') as f:
    lines = f.read().splitlines()
    for line in lines:
        row = [line.lstrip().rstrip(), '']
        worksheet.append(row)



# 保存Excel文件
workbook.save(r'e:\作者.xlsx')