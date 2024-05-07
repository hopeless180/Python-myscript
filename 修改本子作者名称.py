import shutil
import os
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook, Workbook
import time

PROXY = "localhost:7890"
cookies_ = [
    {
        "domain": ".exhentai.org",
        # "expirationDate": 1720655305.508556,
        # "hostOnly": False,
        # "httpOnly": False,
        "name": "igneous",
        "path": "/",
        # "sameSite": "unspecified",
        # "secure": False,
        # "session": False,
        # "storeId": "0",
        "value": "ba4be986b",
        "id": 1,
    },
    {
        "domain": ".exhentai.org",
        # "expirationDate": 1720655305.12921,
        # "hostOnly": False,
        # "httpOnly": False,
        "name": "ipb_member_id",
        "path": "/",
        # "sameSite": "unspecified",
        # "secure": False,
        # "session": False,
        # "storeId": "0",
        "value": "1788404",
        "id": 2,
    },
    {
        "domain": ".exhentai.org",
        # "expirationDate": 1720655305.12924,
        # "hostOnly": False,
        # "httpOnly": False,
        "name": "ipb_pass_hash",
        "path": "/",
        # "sameSite": "unspecified",
        # "secure": False,
        # "session": False,
        # "storeId": "0",
        "value": "21e4f7ba30f6f8b00d4b2dc96549514f",
        "id": 3,
    },
    {
        "domain": ".exhentai.org",
        # "expirationDate": 1689162505.12925,
        # "hostOnly": False,
        # "httpOnly": False,
        "name": "yay",
        "path": "/",
        # "sameSite": "unspecified",
        # "secure": False,
        # "session": False,
        # "storeId": "0",
        "value": "louder",
        "id": 4,
    },
]


def initialize_browser(proxy=None, headless=True):
    chrome_options = webdriver.ChromeOptions()
    if proxy:
        chrome_options.add_argument(f"--proxy-server={proxy}")
    if headless:
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--blink-settings=imagesEnabled=false')
    browser = webdriver.Chrome(options=chrome_options)
    return browser

def process_excel(browser, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        print(f"调查 {row[0]} 中~")
        if (len(row) == 6 and (row[3] or row[4] or row[5])) or row[2] == "错误":
            continue
        time.sleep(30)
        author1, author2, author3 = parse_author_info(search_author(browser, row[0]))
        if author1 is not None:
            ws.cell(row=idx, column=4, value=author1)
            ws.cell(row=idx, column=5, value=author2)
            ws.cell(row=idx, column=6, value=author3)
        else:
            ws.cell(row=idx, column=3, value='错误')
        wb.save(excel_path)
        

def parse_author_info(soup):
    if not soup:
        return None, None, None
    possibility_author = {}
    lines = soup.find(class_='itg gltc').find_all('tr')
    if len(lines) < 2 or lines[-1].text == "No unfiltered results found.":
        return None, None, None
    for item in lines:
        title = item.find(class_="glink")
        if not title:
            continue
        title = title.text
        href = item.find(class_="gl3c glname").find("a").get('href')
        match = re.search(r'\[(.*?)\]', title)
        if not match:
            continue
        content_inside_brackets = match.group(1)
        author = re.search(r'\((.*?)\)', content_inside_brackets).group(1) if '(' in content_inside_brackets else content_inside_brackets
        if author not in possibility_author:
            possibility_author[author] = {
                "main": author,
                "sub": re.sub(r'\(.*?\)', '', content_inside_brackets),
                "count": 1,
                "href": href
            }
        else:
            possibility_author[author]['count'] += 1
    most_possible_author = max(possibility_author.items(), key=lambda x: x[1]['count'])[0]
    author1, author2 = goto_galary_page(browser, possibility_author[most_possible_author]['href'])
    author3 = ",".join(possibility_author.keys())
    return author1, author2, author3
    
def goto_galary_page(browser, href):
    js = f"window.open('{href}')"
    browser.execute_script(js)
    browser.switch_to.window(browser.window_handles[-1])
    print("跳转到详情页")
    try:
        WebDriverWait(browser, 20).until(
            EC.presence_of_element_located(
                (By.ID, "gj")
            )
        )
        WebDriverWait(browser, 20).until(
            EC.presence_of_element_located(
                (By.ID, "gn")
            )
        )
    except TimeoutException as e:
        print(f"发生异常：{e}")
    soup = BeautifulSoup(browser.page_source, 'lxml')
    if soup.find(id='gj') and soup.find(id='gj').text:
        title = soup.find(id="gj").text
    elif soup.find(id='gn') and soup.find(id='gn').text:
        title = soup.find(id='gn').text
    else:
        title = None
    match = re.search(r'\[(.*?)\]', title)
    content_inside_brackets = match.group(1)
    match = re.search(r'\((.*?)\)', content_inside_brackets)
    author_sub = ""
    author = content_inside_brackets
    if match:
        author = match.group(1)
        author_sub = re.sub(r'\(.*?\)', '', content_inside_brackets)
    browser.close()
    browser.switch_to.window(browser.window_handles[-1])
    return author, author_sub

def search_author(broswer, keyword):
    try:
        Input = WebDriverWait(browser, 20).until(
            EC.presence_of_element_located(
                (
                    By.XPATH, 
                    "/html/body/div[2]/div[1]/div[1]/form/div[1]/input[1]"
                )
            )
        )
        Submit = WebDriverWait(browser, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    '/html/body/div[2]/div[1]/div[1]/form/div[1]/input[2]',
                )
            )
        )
        Input.clear()
        Input.send_keys(keyword)
        Submit.click()
        WebDriverWait(browser, 20).until(
            EC.presence_of_element_located(
                (
                    By.XPATH, 
                    "//table[@class='itg gltc']/tbody/tr[1]"
                )
            )
        )
        soup = BeautifulSoup(browser.page_source, 'lxml')
    except:
        print(keyword, "错误")
        browser.save_screenshot(f'keyword_screenshot_{keyword}_error.png')
        return None
    return soup

if __name__ == "__main__":
    browser = initialize_browser('localhost:7890')
    browser.get("https://exhentai.org")
    browser.delete_all_cookies()
    for cookie in cookies_:
        browser.add_cookie(cookie)
    browser.refresh()
    print("==================================")
    print("登录ex绅士")
    print("==================================")
    process_excel(browser, "C:\\Users\\neun\\Desktop\\作者.xlsx")
    browser.quit()

    
